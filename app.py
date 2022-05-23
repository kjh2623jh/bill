import sys
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Protection
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtGui import QIcon, QFont


class BillApp(QWidget):
    def __init__(self):
        super().__init__()
        self.message_log = ''

        self.initUI()

    def initUI(self):
        QToolTip.setFont(QFont('SansSerif', 10))


        getTextButton = QPushButton('내역 넣기', self)
        getTextButton.clicked.connect(self.getMessageLog)

        self.table = QTextBrowser()

        okButton = QPushButton('완료')
        okButton.clicked.connect(self.input_text)

        cancelButton = QPushButton('취소')
        cancelButton.clicked.connect(QCoreApplication.instance().quit)


        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(getTextButton)
        hbox.addStretch(1)

        hbox2 = QHBoxLayout()
        hbox2.addStretch(1)
        hbox2.addWidget(okButton)
        hbox2.addWidget(cancelButton)
        hbox2.addStretch(1)

        vbox = QVBoxLayout()
        vbox.addWidget(self.table)
        vbox.addLayout(hbox)
        vbox.addStretch(3)
        vbox.addLayout(hbox2)
        vbox.addStretch(1)

        self.setLayout(vbox)


        self.setWindowTitle('금액 청구')
        self.setWindowIcon(QIcon('ico/bill.ico'))
        self.resize(500, 350)
        self.center()

        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
    
    def getMessageLog(self):
        text, ok = QInputDialog.getMultiLineText(self, '내역', '메시지 붙여넣기:')

        if ok:
            self.table.setText(str(text))
            self.message_log=str(text)
    
    def input_text(self):
        log = self.message_log.split('[Web발신]\n')
        log_list=[]
        for idx in range(1,len(log)):
            txt = log[idx].split()
            if txt[0] == '농협':
                if txt[1][:2]=="입금": continue
                log_list.append(f"{txt[2]} | {txt[5]} | {txt[1][2:-1]}")
            else:
                txt2 = txt[3].split('(')
                log_list.append(f"{txt[0][2:]} | {txt2[1][:-1]} | {txt2[0][4:-1]}")
        
        self.ca = checkApp(self, log_list)
        self.ca.exec()
        BillApp.make_xlsx(log_list, self.ca.check_idx)
        self.close()
    

    def make_xlsx(log, arr):
        wb=Workbook()
        ws=wb.active
        ws.title="Bill"

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin") \
                            , top=Side(style="thin"), bottom=Side(style="thin"))
        sky_blue = PatternFill(fgColor="B4C6E7", fill_type="solid")
        gray = PatternFill(fgColor="D0CECE", fill_type="solid")

        ws.cell(1,1,"날짜").fill=sky_blue
        ws.cell(1,2).fill=sky_blue
        ws.cell(1,3,"가격").fill=sky_blue
        ws.cell(1,3).alignment=Alignment(horizontal="center")
        for idx in range(len(arr)):
            txt = log[arr[idx]].split(" | ")
            d=txt[0].split('/')
            ws.cell(idx+2,1,f"{d[0]}월 {d[1]}일")
            ws.cell(idx+2,2,txt[1])
            ws[f"C{idx+2}"].value = int(txt[2].replace(',',''))
            # ws.cell(idx+2,3,int(txt[2].replace(',',''))).alignment=Alignment(horizontal="right")
        ws.cell(idx+3,1,"합계").fill=gray
        ws.cell(idx+3,2,"-").fill=gray
        ws.cell(idx+3,3,f"=SUM(C2:C{idx+2})").fill=gray
        ws.column_dimensions['B'].width = 20
        for row in ['A','B','C']:
            for cell in ws[row]:
                if row!='C': cell.alignment=Alignment(horizontal="center")
                cell.border=thin_border

        year= 2022
        month= d[0]
        wb.save(f"xlsx/{year} - {month}.xlsx")
        print('done')



class checkApp(QDialog):
    def __init__(self, parent, loglist):
        super(checkApp, self).__init__(parent)
        self.len_ = len(loglist)
        self.check_list = [None]*self.len_
        self.check_idx = []
        self.log_list = loglist

        self.initUI()

    def initUI(self):
        # self.widget = QWidget()
        for i in range(self.len_):
            self.check_list[i] = QCheckBox(self.log_list[i], self)
            self.check_list[i].toggle()
            self.check_list[i].move(20, 20*i+1)

        okButton = QPushButton('완료')
        okButton.clicked.connect(self.accept)

        cancelButton = QPushButton('취소')
        cancelButton.clicked.connect(self.back)

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(okButton)
        hbox.addWidget(cancelButton)
        hbox.addStretch(1)

        vbox = QVBoxLayout()
        for widget in self.check_list:
            vbox.addWidget(widget)
        vbox.addStretch(1)
        vbox.addLayout(hbox)

        self.setLayout(vbox)

        # scroll = QScrollArea()
        # scroll.setWidget(self.widget)

        # hbox = QHBoxLayout()
        # hbox.addWidget(scroll)

        # hbox2 = QHBoxLayout()
        # hbox2.addStretch(1)
        # hbox2.addWidget(okButton)
        # hbox2.addWidget(cancelButton)
        # hbox2.addStretch(1)

        # vbox = QVBoxLayout()
        # # vbox.addLayout(hbox)
        # vbox.addStretch(9)
        # vbox.addLayout(hbox2)
        # vbox.addStretch(1)

        # self.setLayout(vbox)

        self.setWindowTitle('확인')
        self.setWindowIcon(QIcon('ico/checkbox.ico'))
        self.resize(400, 20*i+40)
        self.center()

        self.show()
    
    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
    
    def accept(self):
        self.check_idx = [i for i in range(self.len_) if self.check_list[i].isChecked()]
        self.close()

    def back(self):
        self.close()


if __name__ == '__main__':
   app = QApplication(sys.argv)
   program = BillApp()
   sys.exit(app.exec_())